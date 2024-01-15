from openpyxl import load_workbook

#https://letscode.com.br/blog/aprenda-a-integrar-python-e-excel
#https://openpyxl.readthedocs.io/en/stable/tutorial.html

#caminho = 'AULAS-EAD-REGISTRO-DE-ATIVIDADE.xlsx'
caminho = 'Aprendendoalerplanilhas.xlsx'
arquivo_excel = load_workbook(caminho, data_only=True)

planilha = arquivo_excel.active
#conteudo = planilha['B5']
#celula = []
#celula.append(conteudo.value)

# planilha.cell(row=7, column=4, value='')
# arquivo_excel.save("Aprendendoalerplanilhas.xlsx")
alunos = []
for i in range(6, 16):
    conteudo = planilha[f'B{i}']
    alunos.append(conteudo.value)

resultadoAlunosSim = {}
resultadoAlunosNao = {}

for a in alunos:
    somasim = 0
    somanao = 0
    for i in range(6, 45):
        celulaAluno = planilha[f'B{i}']
        celulaResposta = planilha[f'C{i}']
        if celulaResposta.value != None:
            if celulaAluno.value == a and celulaResposta.value == 'SIM':
                somasim += 1
                resultadoAlunosSim[a] = somasim
            elif celulaAluno.value == a and (celulaResposta.value == 'NÃO' or celulaResposta.value == 'NAO'):
                somanao += 1
                resultadoAlunosNao[a] = somanao
print("SIM", resultadoAlunosSim)
print("NÃO", resultadoAlunosNao)


indice = 0
for z in range(49, 49 + len(alunos)):
        planilha.cell(row=z, column=5, value=alunos[indice])
        indice += 1

arquivo_excel.save("Aprendendoalerplanilhas.xlsx")

indice = 0
for z in range(49, 49 + len(alunos)):
    if alunos[indice] == planilha[f'E{z}'].value:
        planilha.cell(row=z, column=6, value=resultadoAlunosSim.get(alunos[indice]))
        planilha.cell(row=z, column=7, value=resultadoAlunosNao.get(alunos[indice]))
        indice += 1

arquivo_excel.save("Aprendendoalerplanilhas.xlsx")
























'''
resultadoAlunossim = {}
resultadoAlunosnao = {}


for a in alunos:
    somasim = 0
    somanao = 0
    for i in range(7,28):
        celulaaluno = planilha[f'B{i}']
        celularesposta = planilha[f'C{i}']
        if celulaaluno.value == a and celularesposta.value == 'SIM':
            somasim += 1
            resultadoAlunossim[a] = somasim
        elif celulaaluno.value == a and celularesposta.value == 'NÃO':
            somanao += 1
            resultadoAlunosnao[a] = somanao


print('SIM', resultadoAlunossim)
print('NÃO', resultadoAlunosnao)
'''


'''
caminho = 'Aprendendoalerplanilhas.xlsx'
arquivo_excel = load_workbook(caminho, data_only=True)
planilha = arquivo_excel.active

lista_presentes = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
lista_faltosos = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

def encontrar_posicoes_aluno1():
    posicoes_aluno1 = []

    for i in range (1, 80):
        if planilha["B{}".format(i)].value == "Aluno 1":
            posicoes_aluno1.append(i)
    return posicoes_aluno1

#posicoes_aluno1 = [6, 21, 35]

def encontrar_presencas_e_faltas (lista_presentes, lista_faltosos):
    k = 0

    while k < 3:
        i = 0
        j = posicoes_aluno1[k]

        while (i < 10):

            if planilha["C{}".format(j+i)].value == "SIM":
                lista_presentes[i] +=1
            else:
                lista_faltosos[i] += 1
            i += 1
        k +=1

    return lista_presentes, lista_faltosos

def adicionar_presencas_e_faltas_na_tabela():
    k = 0
    for i in range(49, 59):
        planilha["F{}".format(i)] = lista_presentes[k]
        planilha["G{}".format(i)] = lista_faltosos[k]
        k += 1

    arquivo_excel.save("Aprendendoalerplanilhas.xlsx")


posicoes_aluno1 = encontrar_posicoes_aluno1()

lista_presentes, lista_faltosos = encontrar_presencas_e_faltas(lista_presentes, lista_faltosos)

adicionar_presencas_e_faltas_na_tabela()
'''

